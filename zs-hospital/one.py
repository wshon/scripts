#!/usr/bin/env python3
# coding:utf-8 -*-
"""
@author: SamWang
@project: zs-hospital
@file: one.py
@time: 2019/6/24 17:17
@desc: 
"""
import collections
import hashlib
import json
import os
import pathlib
import time
from datetime import datetime

import openpyxl
import requests
from prompt_toolkit import prompt

FLAG = "gdz"
urlbase = "http://m.zs-hospital.sh.cn"
urlbase_version = "v110/api"
urlbase_zs_login = f'{urlbase}:8065/zslogin/{urlbase_version}'
urlbase_zs_app = f'{urlbase}:8063/zsapp/{urlbase_version}'
urlbase_zs_book = f'{urlbase}:8064/zsbook/{urlbase_version}'

url_login = urlbase_zs_login + "/login/login.action"
url_getScanReport = urlbase_zs_app + "/checkReport/getScanReport.action"
url_findVisitsRecord = urlbase_zs_app + "/visitsRecord/findVisitsRecord.action"
url_getUserStatus = urlbase_zs_app + "/user/getUserStatus.action"


class Unit:
    from Crypto.Cipher import AES
    aes_key = '2015-AES-KEY-HIS'
    aes = AES.new(str.encode(aes_key), AES.MODE_ECB)

    @staticmethod
    def decrypt(src):
        import base64
        return base64.b64decode(Unit.aes.decrypt(src))

    @staticmethod
    def md5(src):
        m5 = hashlib.md5()
        m5.update(src.encode())
        return m5.hexdigest()


class HospitalAPI:

    def __init__(self):
        self.session = requests.Session()

    def login(self, mobile, password):
        cmd = 'login'
        password = Unit.md5(f'{mobile}{password}{FLAG}')
        imei = '00000000-12f5-d045-214f-f3234d6a9e12'
        client_type = 1
        sign = Unit.md5(f'{cmd}clientType={client_type}imie={imei}mobile={mobile}password={password}{FLAG}')
        return self.session.post(url_login, {
            'clientType': client_type,
            'imie': imei,
            'mobile': mobile,
            'password': password,
            'sign': sign
        })

    def run_cmd(self, url, name, data):
        sign_str = name + ''.join([f'{k}={v}' for k, v in data.items()]) + FLAG
        # sign = Unit.md5(f'{cmd}cardNo={self.card_no}token={self.token}type={f_type}{FLAG}')
        data['sign'] = Unit.md5(sign_str)
        return self.session.post(url, data)


class Hospital:
    time = 0
    token = None
    phone = None
    password = None
    card_no = None

    def __init__(self, phone, password, card_no, auto_login=True):
        if os.path.exists('token'):
            with open('token', encoding='utf-8') as f:
                content = f.read()
            if ',' in content:
                self.time, self.token = content.split(',')
                self.time = int(self.time)
                print('自动登陆成功')

        self.api = HospitalAPI()
        self.phone = phone
        self.password = password
        self.card_no = card_no
        if auto_login and not self.valid:
            self.login()

    @property
    def valid(self):
        return time.time() - self.time < 60 * 60

    def login(self):
        res = self.api.login(self.phone, self.password)
        if res.status_code != 200:
            prompt('remote server is error')
        content = json.loads(Unit.decrypt(res.content))
        print(content)
        if content['code'] != '0':
            print(content['msg'])
        self.token = content['data']['token']
        with open('token', 'w', encoding='utf-8') as f:
            f.write(f'{time.time()},{self.token}')
        print('登陆成功')

    def check_login(self):
        data = collections.OrderedDict()
        data['token'] = self.token
        return self.api.run_cmd(url_getUserStatus, 'getUserStatus', data)

    def get_scan_report(self):
        data = collections.OrderedDict()
        data['cardNo'] = self.card_no
        data['token'] = self.token
        data['type'] = 0
        return self.api.run_cmd(url_getScanReport, 'getScanReport', data)

    def get_scan_report_by_scan(self, begin_date):
        data = collections.OrderedDict()
        data['beginDate'] = begin_date
        data['cardNo'] = self.card_no
        data['token'] = self.token
        data['type'] = 1
        return self.api.run_cmd(url_getScanReport, 'getScanReport', data)

    def save_scan_report(self, path='.'):
        res = self.get_scan_report()
        content = Unit.decrypt(res.content)
        data = json.loads(content)
        work_book = openpyxl.Workbook()
        for checkReport in data['data']:
            print(checkReport)
            sheet = work_book.create_sheet(checkReport['bk'])
            row = 1
            for index, head in enumerate(['项目', '类型', '值', '参考值', '单位', '标志', ]):
                sheet.cell(row, index + 1, head)
            row += 1
            for checkReportDto in checkReport['checkReportDtoList']:
                print(checkReportDto)
                sheet.cell(row, 1, checkReportDto['invkind'].strip())
                sheet.cell(row, 2, checkReportDto['itemna'].strip())
                sheet.cell(row, 3, checkReportDto['pvalue'].strip())
                sheet.cell(row, 4, checkReportDto['refrence'].strip())
                sheet.cell(row, 5, checkReportDto['unit'].strip())
                sheet.cell(row, 6, checkReportDto['lmtflag'].strip())
                row += 1
        filename = f'checkReportDtoList_{datetime.now().strftime("%Y%m%d%H%M%S")}.xlsx'
        work_book.save(pathlib.Path(path) / filename)
        return filename


if __name__ == '__main__':
    h = Hospital('', '', '')
    h.save_scan_report()
